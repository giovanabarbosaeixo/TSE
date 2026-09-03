"""
Casamento de "nomes de interesse dos clientes" (planilhas Eixo) contra o
cadastro de parlamentares em exercício — pra sinalizar, na tela de
apuração, quem interessa a qual cliente.

DUAS ARMADILHAS JÁ CUSTARAM TEMPO (histórico, não repetir):

1. A planilha renomeia cabeçalho com o uso — já mudou no meio de uma
   sessão. NUNCA ler uma coluna por um nome fixo só. `campo()` abaixo lê
   por uma LISTA de nomes aceitos (`CAMPOS`) e explode em erro se nenhum
   bater — antes disso, uma coluna renomeada devolvia vazio em silêncio
   (sem exceção) e todo mundo caía na categoria errada sem ninguém notar.

2. Nome de urna não é confiável sozinho:
   - Falso positivo: o senador Cleitinho é CLEITINHO AZEVEDO, e existe um
     Cleitinho estadual (deputado estadual, MDB) diferente. Resolvido
     filtrando por CARGO antes de casar o nome — os dois nunca disputam
     o mesmo cargo.
   - Falso negativo: HUGO, ELMAR, GLEISI, LINDBERGH — apelidos/primeiro
     nome que não batem com o nome completo no cadastro. Resolvido
     casando pelo PRIMEIRO TOKEN do nome completo além do nome exato.
   - Se depois do filtro por cargo/partido/UF ainda sobrar mais de um
     candidato batendo, é ambíguo — desempate por data de nascimento
     quando disponível; sem isso, ou sem batida nenhuma, a linha cai na
     lista de PENDÊNCIAS em vez de adivinhar (adivinhar errado é pior
     que não preencher).
"""

from __future__ import annotations

import io
import re
import unicodedata
from dataclasses import dataclass, field

import pandas as pd
import requests

_HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; PainelTSE/1.0)"}

# Planilha "Recandidaturas 2026" — 513 deputados federais, 81 senadores,
# 1.057 estaduais (abas "Competitividade X (em exercício)") + as listas de
# interesse de cada cliente (abas "Clientes X").
PLANILHA_ID = "1PuvLTeBaHK9uOWC-iZo9T06mgeU7CETcZ-XJLX2Jdfw"

# cargo lógico (bate com CARGOS de tse_api.py) -> nome EXATO da aba na
# planilha. Buscar por nome de aba (parâmetro "sheet=" do gviz), não por
# gid — não precisa descobrir/guardar número de gid nenhum, e sobrevive a
# reordenar abas (só quebra se renomearem a aba em si, que aí quem chama
# recebe HTTP 400 e sabe na hora).
ABAS_ROSTER = {
    "Deputado Federal": "Competitividade Câmara (em exercício)",
    "Senador": "Competitividade Senado (em exercício)",
    "Deputado Estadual": "Competitividade Assembleias (em exercício)",
}
ABAS_CLIENTES = ["Clientes Educação", "Clientes Saúde", "Clientes Esporte"]

# ─── Leitura tolerante a renomeação de cabeçalho ────────────────────────────

# campo lógico -> lista de nomes de cabeçalho já vistos nesta planilha.
# Adicione uma variante aqui (não troque a lógica de leitura) quando a
# planilha for renomeada de novo.
CAMPOS = {
    "cliente": ["Cliente"],
    "nome_interesse": ["Nome do Ator de Interesse", "Nome do Ator"],
    "cargo_partido_uf": ["Cargo / Partido / UF", "Cargo/Partido/UF"],
    "alinhamento": ["Alinhamento (Aliado/Neutro/Opositor)", "Alinhamento"],
    "subtema": ["Subtema / Pauta Específica", "Subtema/Pauta Específica"],
    "relacao": ["Relação (Ex: Frente X, Comissão Y)", "Relação"],
    "pls": ["PLs Relacionados"],
    "observacoes": ["Análise / Observações da Eixo", "Observações"],
    # roster (Competitividade X (em exercício))
    "parlamentar": ["Parlamentar"],
    "partido": ["Partido"],
    "uf": ["UF"],
}


class CampoNaoEncontrado(Exception):
    """Nenhum dos nomes aceitos pra esse campo bateu com o cabeçalho real."""


def indice_campos(cabecalho: tuple, campos_esperados: list[str]) -> dict[str, int]:
    """
    Recebe a linha de cabeçalho crua da planilha e a lista de campos
    lógicos que essa leitura PRECISA achar. Devolve {campo_lógico: índice
    da coluna}.

    Casa por igualdade exata depois de normalizar espaço (strip); NÃO
    tenta "parecido" — se a planilha renomeou de um jeito que não está
    em CAMPOS[campo], isso é erro, não adivinhação. Levanta
    CampoNaoEncontrado imediatamente (nunca devolve vazio em silêncio —
    era exatamente isso que fazia todo mundo cair na categoria errada).
    """
    limpo = [(str(c).strip() if c is not None else "") for c in cabecalho]
    resultado: dict[str, int] = {}
    faltando = []
    for campo in campos_esperados:
        aceitos = CAMPOS.get(campo)
        if aceitos is None:
            raise KeyError(f"Campo lógico desconhecido: {campo!r} (não está em CAMPOS)")
        idx = next((i for i, h in enumerate(limpo) if h in aceitos), None)
        if idx is None:
            faltando.append(campo)
        else:
            resultado[campo] = idx
    if faltando:
        raise CampoNaoEncontrado(
            f"Cabeçalho da planilha mudou: não achei coluna pra {faltando} "
            f"(nomes aceitos hoje: {[CAMPOS[c] for c in faltando]}). "
            f"Cabeçalho real agora: {limpo}. "
            f"Adicione a nova variante em CAMPOS antes de tentar de novo — "
            f"não dá pra seguir sem saber qual coluna é qual."
        )
    return resultado


# ─── Normalização e casamento de nome ──────────────────────────────────────

def normalizar(texto: str) -> str:
    """MAIÚSCULA, sem acento, espaços colapsados. Base de qualquer comparação."""
    if not texto:
        return ""
    sem_acento = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", sem_acento).strip().upper()


_CARGO_ALIASES = {
    "SENADOR": "Senador", "SENADORA": "Senador",
    "DEPUTADO": "Deputado Federal", "DEPUTADA": "Deputado Federal",
    "DEPUTADO FEDERAL": "Deputado Federal", "DEPUTADA FEDERAL": "Deputado Federal",
    "DEPUTADO ESTADUAL": "Deputado Estadual", "DEPUTADA ESTADUAL": "Deputado Estadual",
    "DEPUTADO DISTRITAL": "Deputado Estadual", "DEPUTADA DISTRITAL": "Deputado Estadual",
}


def parse_cargo_partido_uf(texto: str) -> tuple[str | None, str | None, str | None]:
    """
    "Senadora - PT/PE" -> ("Senador", "PT", "PE")
    "Deputado - PSB/SP" -> ("Deputado Federal", "PSB", "SP")
    "Ex-parlamentar - PSDB/CE" -> (None, "PSDB", "CE")  # cargo não mapeável,
    fica None de propósito — quem chama decide se isso é pendência.
    """
    if not texto:
        return None, None, None
    partes = texto.split(" - ", 1)
    cargo_txt = normalizar(partes[0]) if partes else ""
    cargo = _CARGO_ALIASES.get(cargo_txt)
    partido = uf = None
    if len(partes) > 1 and "/" in partes[1]:
        pp, _, u = partes[1].rpartition("/")
        partido = pp.strip() or None
        uf = u.strip().upper() or None
    return cargo, partido, uf


@dataclass
class ResultadoCasamento:
    encontrados: list[dict] = field(default_factory=list)
    pendencias: list[dict] = field(default_factory=list)


def casar_nome(
    nome_interesse: str,
    cargo: str | None,
    partido: str | None,
    uf: str | None,
    roster: list[dict],
) -> tuple[dict | None, str]:
    """
    Casa UM "nome de interesse" contra `roster` (lista de dicts com
    'parlamentar', 'partido', 'uf', 'cargo' já normalizados pelo chamador).

    Ordem de tentativa:
    1. Filtra o roster por CARGO primeiro (resolve o falso positivo tipo
       Cleitinho Senador vs Cleitinho Estadual — nunca disputam o mesmo
       cargo, então filtrar por cargo já separa os dois).
    2. Dentro do filtro por cargo, tenta nome EXATO (normalizado).
    3. Se não achou, tenta por PRIMEIRO TOKEN do nome completo do roster
       == nome de interesse (resolve GLEISI -> GLEISI HOFFMANN etc.).
    4. Se sobrar mais de 1 candidato em qualquer uma dessas tentativas,
       tenta desempatar por partido+UF (se dados baterem, fica só 1).
    5. Se ainda ambíguo, ou zero batida, devolve (None, motivo) — motivo
       vai pra lista de pendências, nunca resolve no achismo.

    Devolve (linha_do_roster_ou_None, motivo).
    """
    alvo = normalizar(nome_interesse)
    if not alvo:
        return None, "nome de interesse vazio"

    candidatos = roster
    if cargo:
        candidatos = [r for r in candidatos if r["cargo"] == cargo]
        if not candidatos:
            return None, f"nenhum parlamentar em exercício no cargo {cargo!r}"

    exatos = [r for r in candidatos if r["parlamentar_norm"] == alvo]
    if len(exatos) == 1:
        return exatos[0], "nome exato"

    if not exatos:
        primeiro_token = [
            r for r in candidatos
            if r["parlamentar_norm"].split(" ")[0] == alvo
            or r["parlamentar_norm"] == alvo
        ]
        exatos = primeiro_token

    if len(exatos) > 1 and (partido or uf):
        afinado = [
            r for r in exatos
            if (not partido or normalizar(r["partido"]) == normalizar(partido))
            and (not uf or r["uf"] == uf)
        ]
        if afinado:
            exatos = afinado

    if len(exatos) == 1:
        return exatos[0], "nome exato" if exatos[0]["parlamentar_norm"] == alvo else "primeiro nome/apelido"
    if len(exatos) > 1:
        nomes = [r["parlamentar"] for r in exatos]
        return None, f"ambíguo mesmo depois de cargo/partido/UF: {nomes} — desempatar por data de nascimento"
    return None, "sem batida nenhuma no cadastro em exercício"


# ─── Busca ao vivo na planilha (por nome de aba, não por gid) ──────────────
#
# Testado inicialmente com o endpoint gviz (?tqx=out:csv&sheet=<nome>), que
# dispensa gid. Achado ao validar contra a planilha real: gviz TRUNCA a
# resposta sem avisar quando a aba tem coluna de texto longo por linha —
# a aba "Competitividade Senado (em exercício)" tem uma coluna com um
# parágrafo de metodologia por linha, e voltaram só 15 das 82 linhas, sem
# erro nenhum (mesma classe de falha silenciosa da armadilha nº 1: parece
# ter funcionado, mas devolveu dado incompleto). Troquei pra baixar a
# planilha inteira como .xlsx (sem esse limite, valida a mesma forma que
# testei offline) e ler as abas dali.

def baixar_planilha(planilha_id: str = PLANILHA_ID, timeout: float = 30):
    """Baixa a planilha inteira (todas as abas) como .xlsx e abre com openpyxl."""
    import openpyxl

    url = f"https://docs.google.com/spreadsheets/d/{planilha_id}/export?format=xlsx"
    resp = requests.get(url, headers=_HEADERS, timeout=timeout)
    resp.raise_for_status()
    return openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)


def _achar_aba(wb, nome_aba: str):
    """
    Acha a aba pelo nome completo — ou pelos 31 primeiros caracteres,
    porque o formato .xlsx trunca nome de aba nesse tamanho (limite do
    próprio formato, não da planilha original no Google Sheets).
    """
    if nome_aba in wb.sheetnames:
        return wb[nome_aba]
    alvo = nome_aba[:31]
    candidatos = [n for n in wb.sheetnames if n == alvo or n.startswith(alvo)]
    if len(candidatos) == 1:
        return wb[candidatos[0]]
    raise CampoNaoEncontrado(
        f"Aba {nome_aba!r} não encontrada (nem truncada em 31 chars). "
        f"Abas disponíveis na planilha agora: {wb.sheetnames}. "
        f"A aba pode ter sido renomeada — confira antes de seguir."
    )


def buscar_aba(wb, nome_aba: str) -> list[list[str]]:
    """Lê uma aba já aberta (`wb`, de baixar_planilha()) como lista de linhas."""
    ws = _achar_aba(wb, nome_aba)
    return [
        ["" if v is None else str(v) for v in row]
        for row in ws.iter_rows(values_only=True)
    ]


def carregar_roster(wb) -> list[dict]:
    """Lê as 3 abas 'Competitividade X (em exercício)' e monta o roster único."""
    roster: list[dict] = []
    for cargo, nome_aba in ABAS_ROSTER.items():
        linhas = buscar_aba(wb, nome_aba)
        cabecalho, *dados = linhas
        idx = indice_campos(tuple(cabecalho), ["parlamentar", "partido", "uf"])
        for row in dados:
            nome = row[idx["parlamentar"]] if idx["parlamentar"] < len(row) else ""
            if not nome:
                continue
            roster.append({
                "cargo": cargo,
                "parlamentar": nome,
                "parlamentar_norm": normalizar(nome),
                "partido": row[idx["partido"]] if idx["partido"] < len(row) else "",
                "uf": (row[idx["uf"]] if idx["uf"] < len(row) else "").strip().upper(),
            })
    return roster


def carregar_interesses_clientes(
    planilha_id: str = PLANILHA_ID,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Baixa as abas Clientes X ao vivo, casa cada "nome de interesse" contra
    o roster de parlamentares em exercício, e devolve (encontrados,
    pendencias) como DataFrame.

    encontrados: cliente, nome_interesse, cargo, partido, uf, parlamentar
    (nome oficial batido no roster), motivo (nome exato / apelido).
    pendencias: cliente, nome_interesse, cargo_partido_uf_bruto, motivo —
    pra revisão manual, nunca resolvidas no achismo.
    """
    wb = baixar_planilha(planilha_id)
    roster = carregar_roster(wb)

    encontrados: list[dict] = []
    pendencias: list[dict] = []

    for nome_aba in ABAS_CLIENTES:
        linhas = buscar_aba(wb, nome_aba)
        cabecalho, *dados = linhas
        idx = indice_campos(tuple(cabecalho), ["cliente", "nome_interesse", "cargo_partido_uf"])
        for row in dados:
            def campo(chave):
                i = idx[chave]
                return row[i] if i < len(row) else ""

            nome_interesse = campo("nome_interesse")
            if not nome_interesse:
                continue  # linha de cabeçalho de grupo (só tem o nome completo do cliente)
            cliente = campo("cliente")
            cargo_txt = campo("cargo_partido_uf")
            cargo, partido, uf = parse_cargo_partido_uf(cargo_txt)
            r, motivo = casar_nome(nome_interesse, cargo, partido, uf, roster)
            if r:
                encontrados.append({
                    "cliente": cliente,
                    "nome_interesse": nome_interesse,
                    "cargo": r["cargo"],
                    "partido": r["partido"],
                    "uf": r["uf"],
                    "parlamentar": r["parlamentar"],
                    "motivo": motivo,
                    "pauta": nome_aba.replace("Clientes ", ""),
                })
            else:
                pendencias.append({
                    "cliente": cliente,
                    "nome_interesse": nome_interesse,
                    "cargo_partido_uf_bruto": cargo_txt,
                    "motivo": motivo,
                    "pauta": nome_aba.replace("Clientes ", ""),
                })

    wb.close()
    return pd.DataFrame(encontrados), pd.DataFrame(pendencias)


def marcar_candidatos(
    df_tse: pd.DataFrame, cargo: str, uf: str, encontrados: pd.DataFrame
) -> pd.DataFrame:
    """
    Cruza os candidatos da apuração (df_tse, já filtrados por `cargo`/`uf`)
    com `encontrados` (saída de carregar_interesses_clientes) e devolve
    df_tse com uma coluna "clientes" — lista dos códigos de cliente
    interessados naquele candidato (vazia se nenhum).

    Casa pelo mesmo casar_nome() usado a planilha inteira, só que aqui o
    "roster" é a própria lista de candidatos da apuração — já vem filtrada
    por cargo+UF de fora, então colisão de nome é rara (poucas dezenas de
    candidatos por corrida, não milhares).
    """
    df_tse = df_tse.copy()
    df_tse["clientes"] = [[] for _ in range(len(df_tse))]

    if encontrados.empty:
        return df_tse

    relevantes = encontrados[(encontrados["cargo"] == cargo) & (encontrados["uf"] == uf)]
    if relevantes.empty:
        return df_tse

    roster_local = [
        {
            "cargo": cargo,
            "parlamentar": nome,
            "parlamentar_norm": normalizar(nome),
            "partido": "",
            "uf": uf,
            "_idx": i,
        }
        for i, nome in enumerate(df_tse["nome"])
    ]

    marcas: dict[int, list[str]] = {}
    for _, linha in relevantes.iterrows():
        r, _motivo = casar_nome(linha["parlamentar"], cargo, None, uf, roster_local)
        if r:
            marcas.setdefault(r["_idx"], []).append(linha["cliente"])

    for i, clientes_lista in marcas.items():
        df_tse.at[df_tse.index[i], "clientes"] = sorted(set(clientes_lista))

    return df_tse
